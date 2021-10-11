import openpyxl
from pathlib import Path
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import statistics
import os

sensorData = []
loadCell = []


def make_zeros(number):
    return [0] * number


def movingAverage(list, average):
    moving_averages = []
    numbers = list
    for i in range(average):
        # print(sum(numbers[:i+1]))
        moving_averages.append(sum(numbers[:i+1])/(i+1))
    # print(moving_averages)
    window_size = average
    i = 0

    while i < len(numbers) - window_size + 1:
        this_window = numbers[i: i + window_size]
        window_average = sum(this_window) / window_size
        moving_averages.append(window_average)
        i += 1
    return moving_averages


def calculate_sma(data, days):
    sma = [sum(data[:days]/days)]
    return sma


def calculate_ema(prices, days, smoothing=2):
    ema = [sum(prices[:days]) / days]
    for price in prices[days:]:
        ema.append((price * (smoothing / (1 + days))) +
                   ema[-1] * (1 - (smoothing / (1 + days))))
    return ema


def main():
    setting = []
    myVariable = []
    f = open("setting.txt", "r")
    for x in f:
        setting.append(x)
    value = float
    macd = 1.25
    LMA = 150
    SMA = 30
    setTime = 200  # stand for time start detection
    ColName = 'Encoder'
    path = './'

    for s in setting:

        variable = s
        slice = variable[variable.find('"') + 1:]
        var = slice[:slice.find('"')]
        # print(var)
        if "macd" in s:
            macd = float(var)
        elif "LMA" in s:
            LMA = int(var)
        elif "SMA" in s:
            SMA = int(var)
        elif "setTime" in s:
            setTime = int(var)
        elif "ColName" in s:
            ColName = str(var)
        elif "path" in s:
            path = str(var)

    print(macd, LMA, SMA, setTime, ColName, path)
    excelList = next(os.walk(path))[2]
    print(excelList)
    for n in excelList:
        name = n
        if 'xlsx' in n:
            #  หาค่า moving average  2 ค่า ต่อ 1 sensor  variable = function (interval)
            #  หา mst
            #  plot
            #xlsx_file = Path('./160721')
            #print(xlsx_file)
            wb_obj = openpyxl.load_workbook(path + n)
            sheet = wb_obj.active
            ColNames = {}
            Current = 0
            for COL in sheet.iter_cols(1, sheet.max_column):
                ColNames[COL[0].value] = Current
                Current += 1
            for row_cells in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                try:
                    if isinstance(row_cells[ColNames[ColName]].value, value):
                        sensorData.append(
                            row_cells[ColNames[ColName]].value)
                    # loadCell.append(row_cells[ColNames['Load cell']].value)
                except:
                    print("NO DATA")
            print(sensorData)
            raw_filtered = movingAverage(sensorData, SMA)
            longMA = movingAverage(raw_filtered, LMA)
            shortMA = movingAverage(raw_filtered, SMA)
            print(raw_filtered)
            #eshortMA = calculate_ema(raw_filtered, 30)
            #elongMA = calculate_ema(raw_filtered, 150)
            dif = len(shortMA) - len(longMA)
            shortMA = shortMA[dif:]
            #macd_ema = []
            macd_mv = []
            check = 0
            mst = 0
            #print(len(longMA), len(raw_filtered))
            for n in range(len(longMA)):
                if len(longMA) >= 1000:
                    macd = 1.0
                macd_mv.append(shortMA[n] - longMA[n])
                if macd_mv[n] > macd and check == 0 and n > setTime:
                    #print('n', macd_mv[n])
                    mst = n
                    check = 10

            # fig, ax = plt.plot(2)
            # ax[1].set_ylim(-2, 5)
            # ax[0].plot(sensorData[dif:], color='red')
            plt.figure(figsize=(10, 5))
            #plt.ylim(0, 50)
            #plt.xlim(-10, 900)
            plt.locator_params(axis="x", nbins=20)
            # plt.gca().set_aspect('equal', adjustable='box')
            plt.plot(raw_filtered[dif:], color='green')
            plt.plot(shortMA, color='pink')
            plt.plot(longMA, color='red')
            plt.axvline(x=mst, color='y', lw=2)
            #plt.plot(shortMA, color='red')
            #plt.plot(macd_mv, color='blue')
            plt.legend(['Raw Data', 'MA' + str(SMA), 'MA' +
                       str(LMA), 'MST :' + str(mst)])

            fig = plt.figure(1)
            # ax[1].plot(raw_filter[dif:], color='blue')
            # ax[1].plot(macd_mv[dif:], color='pink')
            plt.savefig(str(name)+'.png')
            # fig.set_size_inches(15, 10)
            # fig.canvas.draw()

            plt.close()

            sensorData.clear()

            # print(len(ema30), len(ma30))


if __name__ == "__main__":
    main()
