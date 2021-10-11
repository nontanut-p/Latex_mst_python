import openpyxl
from pathlib import Path
import numpy as np
import matplotlib.pyplot as plt
import statistics

# add graph 30 percent
humanLaserin = [0, 0, 0, 0, 0, 0, 1248, 656, 1306, 793, 803, 650]
humanLoadin = [621, 614, 778, 1308, 1339, 794, 1258, 662, 1315, 785, 787, 640]
forceOnlaser = []
forceOnLoadcell = []
forceHumanLOAD = []
forceHumanLaser = []
filename = ''
short = 30
long = 150
humanMST = 0
folderName = "L13-L24"
MACDPOINT = 1.25
myname = ".MACD" + str(MACDPOINT) + "MA" + str(short) + \
    "MA" + str(long) + ".png"
timeMT = []
loadCell = []
laserDis = []
laserDisMA = []
loadCellMA = []
shortMovingLaser = []
longMovingLaser = []
shortMovingLoad = []
longMovingLoad = []
MSTLaser = []
MSTLoad = []
MSTLOADLIST = []
MSTLASERLIST = []
MACDLASER = []
MACDLOAD = []
MVSHORT_MACDLASER = []
MVLONG_MACDLASER = []
MVSHORT_MACDLOAD = []
MVLONG_MACDLOAD = []
MACDMACDLASER = []
MACDMACDLOAD = []

stochasticLOAD = []
stochasticLASER = []
stochasticLOADAV = []
stochasticLASERAV = []

zeros = []
ones = []
mstLaser = 0
mstLoad = 0
pToline = []
failStatus = 0
humanLaser = []
humanLoad = []
print(humanLaserin[11])
if len(humanLaserin) == len(humanLoadin):
    print('yes', len(humanLoadin))


def readData():

    global loadCell, laserDis, timeMT
    xlsx_file = Path(folderName, filename+'.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active
    ColNames = {}
    Current = 0
    for COL in sheet.iter_cols(1, sheet.max_column):
        ColNames[COL[0].value] = Current
        Current += 1

    for row_cells in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        try:
            timeMT.append(row_cells[ColNames['MST Time(s)']].value)
            laserDis.append(row_cells[ColNames['Distance(V)']].value)
            loadCell.append(row_cells[ColNames['Load cell (g)']].value)
        except:
            print("NO DATA")

    total = 0
    for num in range(len(laserDis)):
        total += laserDis[num]
        laserDisMA.append(statistics.mean(laserDis[num: (num+short)]))
        loadCellMA.append(statistics.mean(loadCell[num: (num+short)]))

    movingAV()


def movingAV():
    global loadCellMA, laserDisMA, mstLaser, mstLoad, failStatus
    myrange = len(laserDisMA)
    stateLOAD = 1
    stateLASER = 1
    shortlenght = short
    longlenght = long
    # statistics.mean(y1_vals[arraylenght-12:arraylenght])
    for arraylenght in range(myrange):
        if arraylenght <= 1 and arraylenght <= shortlenght:
            shortMovingLoad.append(0)
            longMovingLoad.append(0)
            shortMovingLaser.append(0)
            longMovingLaser.append(0)
            MACDLASER.append(0)
            MACDLOAD.append(0)
        if arraylenght > shortlenght and arraylenght <= longlenght:
            ShortEMA = statistics.mean(
                laserDisMA[arraylenght-shortlenght:arraylenght])
            LongEMA = statistics.mean(
                laserDisMA[arraylenght-shortlenght:arraylenght])
            myMACD = ShortEMA - LongEMA
            MACDLASER.append(0)
            MACDLOAD.append(0)
            shortMovingLoad.append(0)
            longMovingLoad.append(0)
            shortMovingLaser.append(0)
            longMovingLaser.append(0)
        if arraylenght > longlenght and arraylenght < (longlenght + shortlenght):
            ShortEMA = statistics.mean(
                laserDisMA[arraylenght-shortlenght:arraylenght])
            shortMovingLaser.append(ShortEMA)
            LongEMA = statistics.mean(
                laserDisMA[arraylenght-longlenght: arraylenght])
            longMovingLaser.append(LongEMA)
            myMACD = ShortEMA - LongEMA
            MACDLASER.append(0)
            ShortEMA = statistics.mean(
                loadCellMA[arraylenght-shortlenght:arraylenght])
            shortMovingLoad.append(ShortEMA)
            LongEMA = statistics.mean(
                loadCellMA[arraylenght-longlenght: arraylenght])
            longMovingLoad.append(LongEMA)
            myMACD = ShortEMA - LongEMA
            MACDLOAD.append(0)
        if arraylenght >= (longlenght + shortlenght):
            if arraylenght == humanMST:  # MST FIXED
                ShortEMA = statistics.mean(
                    laserDisMA[arraylenght-shortlenght:arraylenght])
                shortMovingLaser.append(ShortEMA)
                LongEMA = statistics.mean(
                    laserDisMA[arraylenght-longlenght: arraylenght])
                longMovingLaser.append(LongEMA)
                myMACD = ShortEMA - LongEMA
                MACDLASER.append(myMACD)
                humanLoad.append(myMACD)
                ShortEMA = statistics.mean(
                    loadCellMA[arraylenght-shortlenght:arraylenght])
                shortMovingLoad.append(ShortEMA)
                LongEMA = statistics.mean(
                    loadCellMA[arraylenght-longlenght: arraylenght])
                longMovingLoad.append(LongEMA)
                myMACD = ShortEMA - LongEMA
                MACDLOAD.append(myMACD)
                humanLaser.append(myMACD)
            else:
                ShortEMA = statistics.mean(
                    laserDisMA[arraylenght-shortlenght:arraylenght])
                shortMovingLaser.append(ShortEMA)
                LongEMA = statistics.mean(
                    laserDisMA[arraylenght-longlenght: arraylenght])
                longMovingLaser.append(LongEMA)
                myMACD = ShortEMA - LongEMA
                MACDLASER.append(myMACD)
                ShortEMA = statistics.mean(
                    loadCellMA[arraylenght-shortlenght:arraylenght])
                shortMovingLoad.append(ShortEMA)
                LongEMA = statistics.mean(
                    loadCellMA[arraylenght-longlenght: arraylenght])
                longMovingLoad.append(LongEMA)
                myMACD = ShortEMA - LongEMA
                MACDLOAD.append(myMACD)

            if stateLASER == 1 or stateLOAD == 1:

                if MACDLOAD[arraylenght] > MACDPOINT and stateLOAD == 1:
                    mstLoad = arraylenght
                    stateLOAD = 0

                if MACDLASER[arraylenght] > MACDPOINT and stateLASER == 1:
                    mstLaser = arraylenght
                    stateLASER = 0

    plotXY()


def plotXY():

    fig, ax = plt.subplots(3)
    global loadCellMA, laserDisMA, timeMT, shortMoving, longMoving, zeros, ones, mstLaser, mstLoad
    # print('stoc',stochasticLASER)

    y_axisloadCell = loadCellMA
    y_axisLaser = laserDisMA
    ax[0].plot(y_axisLaser, color='red')
    ax[0].plot(shortMovingLaser, color='deeppink')
    ax[0].plot(longMovingLaser, color='darkred')
    ax[0].plot(y_axisloadCell, color='green')
    ax[0].plot(shortMovingLoad, color='lime')
    ax[0].plot(longMovingLoad, color='darkgreen')

    MSTLaser.append(mstLaser)
    MSTLoad.append(mstLoad)
    ax[0].legend(["LASER",  "LASER: : MA " + str(short), "LASER :  MA " + str(long), "LOADCELL",
                 "LOADCELL: : MA " + str(short), "LOADCELL :  MA " + str(long)], loc="lower right")

    if failStatus == 1:
        text = "ERROR!!!"
        ax[1].text(500, 0.5, text, color='r', horizontalalignment='center',
                   verticalalignment='center', size=50)
    else:
        text = "MST LASER : " + str(mstLaser)
        ax[1].scatter(mstLaser, MACDPOINT, color='deeppink',
                      marker='o', zorder=10, s=50)
        ax[1].text(mstLaser, 0.5, text, color='deeppink',
                   horizontalalignment='center', verticalalignment='center', size=18)
        ax[1].axvline(x=mstLaser, color='deeppink', lw=2)
        ## LOAD CELL ##
        text = "MST LOAD CELL : " + str(mstLoad)
        ax[2].scatter(mstLoad, MACDPOINT, color='lime',
                      marker='o', zorder=10, s=50)
        ax[2].text(mstLoad, 2.5, text, color='lime',
                   horizontalalignment='center', verticalalignment='center', size=18)
        ax[2].axvline(x=mstLoad, color='lime', lw=2)
        ax[1].axvline(x=humanLaserin[int(filename) - 1], color='y', lw=2)
        ax[2].axvline(x=humanLoadin[int(filename)-1], color='y', lw=2)

    ax[1].set_ylim(-2, 3)
    ax[2].set_ylim(-2, 3)
    ax[1].plot(MACDLASER, color='r')
    ax[1].legend(["LASER ", "HUMAN"], loc="lower right")
    ax[2].plot(MACDLOAD, color='g')
    ax[2].legend(["LOADCELL ", "HUMAN"],
                 loc="lower right")
    ax[1].arrow(0.2, 0.3, 0.7, 0.0, color='green',
                head_length=0.07, head_width=0.025,
                length_includes_head=True)

    ax[1].arrow(0.9, 0.3, -0.7, 0.0, color='green',
                head_length=0.07, head_width=0.025,
                length_includes_head=True)
    ax[1].axhline(y=1, color='r', linestyle='-')
    ax[1].axhline(y=0, color='b', linestyle='-')
    ax[2].axhline(y=1, color='r', linestyle='-')
    ax[2].axhline(y=0, color='b', linestyle='-')
    # ax[1].axhline(y=[])  # fix here
    # ax[1].axhline(y=.5, xmin=0.25, xmax=0.75)
    forceOnlaser.append(laserDisMA[mstLaser])
    forceOnLoadcell.append(loadCellMA[mstLoad])
    try:
        forceHumanLOAD.append(loadCellMA[humanLoadin[int(filename) - 1]])
    except:
        forceHumanLOAD.append('no-data')
    try:
        forceHumanLaser.append(laserDisMA[humanLaserin[int(filename) - 1]])
    except:
        forceHumanLaser.append('no-data')

    fig.set_size_inches(15, 10)
    fig.canvas.draw()
    plt.savefig(filename + myname, dpi=400)
    y_axisloadCell.clear()
    y_axisLaser.clear()
    #print("laser", humanLaser)
    #print('load', humanLoad)
    #print('mstLoad =',  MSTLoad)
    MSTLOADLIST.append(MSTLoad)
    MSTLASERLIST.append(MSTLaser)
    #print('mstLaser =', MSTLaser)
    plt.show()


def repeat():
    global filename, loadCell, failStatus, timeMT, mstLoad, mstLaser, zeros, ones, mst, shortMovingLaser, longMovingLaser, shortMovingLoad, longMovingLoad, MACDLASER, MACDLOAD
    MSTLaser.clear()
    MSTLoad.clear()
    for i in range(1, 2):  # filenumber
        filename = str(i)
        readData()
        timeMT.clear()
        loadCell.clear()
        laserDis.clear()
        laserDisMA.clear()
        loadCellMA.clear()
        shortMovingLaser.clear()
        longMovingLaser.clear()
        shortMovingLoad.clear()
        longMovingLoad.clear()
        MACDLASER.clear()
        MACDLOAD.clear()
        zeros.clear()
        ones.clear()
        stochasticLASER.clear()
        stochasticLOAD.clear()
        stochasticLOADAV.clear()
        stochasticLASERAV.clear()
        mstLaser = 0
        mstLoad = 0
        failStatus = 0
        print(laserDis)
        MVSHORT_MACDLASER.clear()
        MVLONG_MACDLASER.clear()
        MVSHORT_MACDLOAD.clear()
        MVLONG_MACDLOAD.clear()
        MACDMACDLASER.clear()
        MACDMACDLOAD .clear()
        plt.clf()
    print('FORCE ON LOAD ALGO', forceOnLoadcell)
    print('Human LOADCELL ', forceHumanLOAD)
    print('Force on laser ALGO', forceOnlaser)
    print('Force on laserHUMAN', forceHumanLaser)


# plotxy()
# plotdiff()
# plotprojectdiff()
repeat()
# readData()
